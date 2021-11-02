import os
import argparse
from decimal import Decimal
from openpyxl import load_workbook

import util
import statusinvest

def main(statusinvest_template_path: str, b3_export_path: str):
  wb_statusinvest = load_workbook(statusinvest_template_path)
  ws_statusinvest = wb_statusinvest.worksheets[0]
  ws_b3 = load_workbook(b3_export_path).worksheets[0]

  b3_rows = [[y.value for y in x] for x in ws_b3.iter_rows(min_row=2)]
  b3_rows = list(reversed(b3_rows))

  orders = [{
    'date': row[0],
    'category': statusinvest.get_asset_category(row[5]),
    'ticker': util.normalize_ticker(row[5]),
    'side': row[1][0],
    'quantity': row[6],
    'price': row[7],
    'broker': statusinvest.BROKER,
    'fee': util.get_total_fee(row[6] * Decimal(str(row[7]))),
  } for row in b3_rows]

  for row_pos in range(2, len(orders) + 1):
    index = row_pos - 2
    order = orders[index]
    ws_statusinvest[f'A{row_pos}'] = order['date']
    ws_statusinvest[f'B{row_pos}'] = order['category']
    ws_statusinvest[f'C{row_pos}'] = order['ticker']
    ws_statusinvest[f'D{row_pos}'] = order['side']
    ws_statusinvest[f'E{row_pos}'] = order['quantity']
    ws_statusinvest[f'F{row_pos}'] = order['price']
    ws_statusinvest[f'G{row_pos}'] = order['broker']
    ws_statusinvest[f'I{row_pos}'] = order['fee']

  wb_statusinvest.save('statusinvest.xlsx')

if __name__ == '__main__':
  parser = argparse.ArgumentParser()
  parser.add_argument('-s', dest='statusinvest_template', help='StatusInvest spreadsheet template')
  parser.add_argument('-b', dest='b3_export', help='Spreadsheet export from B3')
  args = parser.parse_args()

  main(
    os.path.join(os.getcwd(), args.statusinvest_template),
    os.path.join(os.getcwd(), args.b3_export))
