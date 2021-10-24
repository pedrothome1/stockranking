import re
from operator import itemgetter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
import data.investsite as investsite
import data.fundamentus as fundamentus

def setup_sheet(ws: Worksheet):
  ws.column_dimensions['A'].width = 20
  ws.column_dimensions['B'].width = 20
  ws.column_dimensions['C'].width = 20

  ws['A1'] = 'Papel'
  ws['A1'].font = Font(name='Calibri', bold=True, color='ffffff')
  ws['A1'].fill = PatternFill(start_color='8e86ae', end_color='8e86ae', fill_type='solid')
  ws['B1'] = 'EV/EBIT'
  ws['B1'].font = Font(name='Calibri', bold=True, color='ffffff')
  ws['B1'].fill = PatternFill(start_color='8e86ae', end_color='8e86ae', fill_type='solid')
  ws['C1'] = 'Liquidez'
  ws['C1'].font = Font(name='Calibri', bold=True, color='ffffff')
  ws['C1'].fill = PatternFill(start_color='8e86ae', end_color='8e86ae', fill_type='solid')

def fill_sheet(ws: Worksheet, stocks: list[dict], banks: list[str], insurance_companies: list[str]):
  setup_sheet(ws)

  excepts = []
  excepts.extend(banks)
  excepts.extend(insurance_companies)

  MIN_LIQUIDITY = 200_000

  stock_list = sorted([
    x for x in stocks
    if x['ticker'] not in excepts and
    not re.search(r'3[2-5]$', x['ticker']) and
    x['volume'] >= MIN_LIQUIDITY and
    x['evebit'] != 'NA' and
    x['ebit_margin'] != 'NA' and
    x['ebit_margin'] >= 0
  ], key=itemgetter('company', 'volume'))

  # When there is more than one ticker of the same company,
  # keep the one with the highest liquidity (the list must be sorted first)
  # 
  # Example:
  # ['AAAA', 'AAAA', 'AAAA', 'BBBB', 'BBBB']
  # ['AAAA', 'AAAA', 'BBBB', 'BBBB']
  # ['AAAA', 'BBBB', 'BBBB']
  # ['AAAA', 'BBBB']
  i = 0
  while True:
    if i >= len(stock_list):
      break
    while i + 1 < len(stock_list) and stock_list[i]['company'] == stock_list[i + 1]['company']:
      del stock_list[i]
    i += 1

  for i in range(0, len(stock_list)):
    ws[f'A{i + 2}'] = stock_list[i]['ticker']
    ws[f'B{i + 2}'] = stock_list[i]['evebit']
    ws[f'C{i + 2}'] = stock_list[i]['volume']

  ws.auto_filter.ref = ws.dimensions

def main():
  wb = Workbook()

  wb.remove(wb.active)
  wb.create_sheet('Fundamentus')
  wb.create_sheet('InvestSite')

  fill_sheet(
    wb.worksheets[0],
    fundamentus.get_stocks(),
    fundamentus.get_banks_tickers(),
    fundamentus.get_insurance_companies_tickers())
  fill_sheet(
    wb.worksheets[1],
    investsite.get_stocks(),
    investsite.get_banks_tickers(),
    investsite.get_insurance_companies_tickers())

  wb.save('output.xlsx')

if __name__ == '__main__':
  main()
